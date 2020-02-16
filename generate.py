from openpyxl import load_workbook
from jinja2 import Template
from .secrets import *

with open('markers.template.html', 'r') as stream:
    template_text = stream.read()

with open('index.template.html', 'r') as stream:
    index_template = stream.read()

wb = load_workbook(filename="data.xlsx")

sheet = wb.active
items = []
for row in sheet.iter_rows():
    cells = list(row)
    if cells[0].value is None:
        break

    d = {
        'text': cells[0].value.replace("\"", "'"),
        'location': [
            cells[2].value,
            cells[3].value,
        ],
        'info': cells[5].value.replace("\"", "'") + "<br/> More info at: <a href=\\\"{}\\\">{}</a>".format(cells[4].value,
                                                                                                       cells[4].value),
    }

    if cells[6].value is not None:
        d['icon'] = "icons/" + cells[6].value
        del d['text']

    items.append(d)


template = Template(template_text)

text = template.render({
    'markers': items,
})

with open('index.html', 'w') as stream:
    stream.write(index_template.replace("MARKERDEFINITIONS", text).replace("MY_API_KEY", API_KEY_local))

with open('public/index.html', 'w') as stream:
    stream.write(index_template.replace("MARKERDEFINITIONS", text).replace("MY_API_KEY", API_KEY_prod))
